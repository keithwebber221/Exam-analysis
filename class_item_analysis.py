"""
class_item_analysis.py
全級班際項目分析模組（第二階段）
依賴：pandas, numpy, openpyxl
整合方式：由 app.py 呼叫 generate_class_analysis_excel()
"""

import io
import re
import zipfile
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════
# 配色常數
# ══════════════════════════════════════════════════════════════
# 班別顏色（自動分配，支援最多 8 班）
PALETTE = [
    ("1F9E9E", "E0F4F4", "0D6E6E"),  # Teal
    ("7D8B99", "EDF0F3", "4A5568"),  # Slate
    ("8E6BB0", "F0E6FA", "5B3A7E"),  # Alexandrite
    ("C0392B", "FDECEA", "922B21"),  # Lava
    ("E67E22", "FEF0E6", "A04000"),  # Amber
    ("27AE60", "E9F7EF", "1A6B3A"),  # Emerald
    ("2980B9", "EAF4FB", "1A5276"),  # Ocean
    ("8E44AD", "F5EEF8", "6C3483"),  # Purple
]

def _build_palette(classes):
    """為班別清單分配顏色，回傳 {cls: (bar_color, bg_color, dark_color)}"""
    out = {}
    for i, cls in enumerate(classes):
        out[cls] = PALETTE[i % len(PALETTE)]
    return out

# ══════════════════════════════════════════════════════════════
# 樣式工具
# ══════════════════════════════════════════════════════════════
def _cfill(hex6):
    return PatternFill("solid", fgColor=hex6)

NO_FILL = PatternFill(fill_type=None)

def _fnt(bold=False, size=11, color="000000", name="Microsoft JhengHei"):
    return Font(bold=bold, size=size, color=color, name=name)

_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LFT = Alignment(horizontal="left",   vertical="center", indent=1)

def _thin_border(color="BBCAD6"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _med_border(color="7F9AB5"):
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _sc(cell, val=None, fill=None, font=None, aln=None, bdr=None, fmt=None):
    if val is not None:
        cell.value = val
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if aln is not None:
        cell.alignment = aln
    if bdr is not None:
        cell.border = bdr
    if fmt is not None:
        cell.number_format = fmt

def _set_merged_border(ws, r1, c1, r2, c2, fill_hex, font_obj, val,
                       fmt="@", bdr=None):
    """合併儲存格並對每個格子正確設框線"""
    if bdr is None:
        bdr = _thin_border()
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    left_s  = bdr.left
    right_s = bdr.right
    top_s   = bdr.top
    bot_s   = bdr.bottom
    none_s  = Side(style=None)
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            cell = ws.cell(r, c)
            cell.fill      = _cfill(fill_hex)
            cell.alignment = _CTR
            l  = left_s  if c == c1 else none_s
            ri = right_s if c == c2 else none_s
            t  = top_s   if r == r1 else none_s
            b  = bot_s   if r == r2 else none_s
            cell.border = Border(left=l, right=ri, top=t, bottom=b)
    main = ws.cell(r1, c1)
    main.value        = val
    main.font         = font_obj
    main.number_format= fmt
    main.alignment    = _CTR

def _make_bar(pct, width=20, fill_char="█", empty_char="░"):
    """等寬文字進度條，pct=0~1"""
    filled = max(0, min(width, round(pct * width)))
    return fill_char * filled + empty_char * (width - filled)

def _diff_fill(d):
    if d >  0.05: return _cfill("C6EFCE")
    if d < -0.05: return _cfill("FFC7CE")
    return _cfill("FFFBEA")

def _diff_color(d):
    if d >  0.05: return "1E6E30"
    if d < -0.05: return "9B1C1C"
    return "7D6608"

# ══════════════════════════════════════════════════════════════
# 統計計算
# ══════════════════════════════════════════════════════════════
def _compute_stats(df, q_cols, max_scores, paper_map, class_col):
    """
    回傳：
      class_stats:  {cls: {q: {mean, mean_pct, std}}}
      grade_stats:  {q: {mean, mean_pct, std}}
      classes:      sorted list of class labels
      paper_totals: {paper: {cls: {mean,std,max,min}, 'grade': ...}}
      combined_totals: {cls: {mean,std,max,min}, 'grade': ...}
    """
    classes = sorted(class_col.unique().tolist())
    df2 = df[q_cols].copy()
    df2["_cls"] = class_col.values

    class_stats = {}
    for cls in classes:
        sub = df2[df2["_cls"] == cls][q_cols]
        class_stats[cls] = {}
        for q in q_cols:
            mx = float(max_scores[q])
            class_stats[cls][q] = {
                "mean":     round(sub[q].mean(), 2),
                "mean_pct": sub[q].mean() / mx if mx > 0 else 0,
                "std":      round(sub[q].std(ddof=1), 2),
            }

    grade_stats = {}
    for q in q_cols:
        mx = float(max_scores[q])
        grade_stats[q] = {
            "mean":     round(df2[q].mean(), 2),
            "mean_pct": df2[q].mean() / mx if mx > 0 else 0,
            "std":      round(df2[q].std(ddof=1), 2),
        }

    papers = sorted(set(paper_map.values()))

    def _paper_total_stats(subset_df, p_cols):
        raw = subset_df[p_cols].sum(axis=1)
        mx  = float(max_scores[p_cols].sum())
        return {"mean": round(raw.mean(),2), "std": round(raw.std(ddof=1),2),
                "max": int(raw.max()), "min": int(raw.min()),
                "max_score": mx,
                "mean_pct": raw.mean()/mx if mx>0 else 0}

    paper_totals = {}
    for p in papers:
        p_cols = [q for q in q_cols if paper_map[q] == p]
        if not p_cols:
            continue
        paper_totals[p] = {}
        for cls in classes:
            sub = df2[df2["_cls"] == cls]
            paper_totals[p][cls] = _paper_total_stats(sub, p_cols)
        paper_totals[p]["grade"] = _paper_total_stats(df2, p_cols)

    all_cols = q_cols
    combined_totals = {}
    for cls in classes:
        sub = df2[df2["_cls"] == cls]
        combined_totals[cls] = _paper_total_stats(sub, all_cols)
    combined_totals["grade"] = _paper_total_stats(df2, all_cols)

    return class_stats, grade_stats, classes, paper_totals, combined_totals

# ══════════════════════════════════════════════════════════════
# 各班分析工作表
# ══════════════════════════════════════════════════════════════
def _make_class_sheet(wb, cls, q_cols, max_scores, paper_map,
                      class_stats, grade_stats, df, class_col, palette):
    bar_hex, bg_hex, dark_hex = palette[cls]
    other_classes = [c for c in sorted(class_col.unique()) if c != cls]
    df_other = df[class_col != cls]

    # 計算其他班統計
    other_stats = {}
    for q in q_cols:
        mx = float(max_scores[q])
        other_stats[q] = {
            "mean":     round(df_other[q].mean(), 2),
            "mean_pct": df_other[q].mean() / mx if mx > 0 else 0,
            "std":      round(df_other[q].std(ddof=1), 2),
        }

    # 差距圖刻度（每卷獨立計算）
    papers = sorted(set(paper_map.values()))
    scale_map = {}
    for p in papers:
        p_qs = [q for q in q_cols if paper_map[q] == p]
        max_d = max(
            (abs(class_stats[cls][q]["mean_pct"] - grade_stats[q]["mean_pct"])
             for q in p_qs), default=0.05
        )
        scale_map[p] = max(0.05, round(max_d + 0.02, 2))

    ws = wb.create_sheet(f"{cls} 分析")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    n_cls   = int((class_col == cls).sum())
    n_other = len(df_other)
    n_all   = len(df)
    other_label = "+".join(other_classes)
    NCOLS = 18

    # ── 標題列 ──
    ws.row_dimensions[2].height = 32
    ws.merge_cells(f"B2:{get_column_letter(1+NCOLS)}2")
    _sc(ws["B2"],
        val=f"{cls}班　項目分析　·　本班 vs 全級　／　本班 vs {other_label}合拼",
        fill=_cfill(dark_hex),
        font=_fnt(bold=True, size=14, color="FFFFFF"),
        aln=_CTR)
    ws.row_dimensions[3].height = 14
    ws.merge_cells(f"B3:{get_column_letter(1+NCOLS)}3")
    _sc(ws["B3"],
        val=f"本班人數：{n_cls}　　{other_label} 合拼人數：{n_other}　　全級人數：{n_all}",
        font=_fnt(size=10, color="595959"), aln=_CTR, fill=NO_FILL)

    # ── 表頭（兩行）──
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 18
    col = 2

    for lbl in ["題號", "滿分"]:
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
        _sc(ws.cell(4, col), val=lbl, fill=_cfill("D9D9D9"),
            font=_fnt(bold=True, size=11, color="1F3864"), aln=_CTR)
        col += 1

    # 本班（5欄）
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+4)
    _sc(ws.cell(4, col), val=f"本班（{cls}）",
        fill=_cfill(dark_hex), font=_fnt(bold=True, size=12, color="FFFFFF"), aln=_CTR)
    for o, lbl in enumerate(["人數", "平均分", "平均%", "平均%圖表", "S.D."]):
        _sc(ws.cell(5, col+o), val=lbl,
            fill=_cfill(bg_hex), font=_fnt(bold=True, size=10, color=dark_hex), aln=_CTR)
    col += 5

    # 全級（4欄）
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+3)
    _sc(ws.cell(4, col), val="全　級",
        fill=_cfill("5D6A77"), font=_fnt(bold=True, size=12, color="FFFFFF"), aln=_CTR)
    for o, lbl in enumerate(["平均分", "平均%", "平均%圖表", "S.D."]):
        _sc(ws.cell(5, col+o), val=lbl,
            fill=_cfill("EAECEE"), font=_fnt(bold=True, size=10, color="3D4E5C"), aln=_CTR)
    col += 4

    # vs 全級（2欄）
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
    _sc(ws.cell(4, col), val="本班 vs 全級",
        fill=_cfill("2C3E50"), font=_fnt(bold=True, size=12, color="FFFFFF"), aln=_CTR)
    for o, lbl in enumerate(["差距%", "差距圖表"]):
        _sc(ws.cell(5, col+o), val=lbl,
            fill=_cfill("D5F5E3"), font=_fnt(bold=True, size=10, color="1E8449"), aln=_CTR)
    col += 2

    # 其他班（4欄）
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+3)
    _sc(ws.cell(4, col), val=f"{other_label} 合拼",
        fill=_cfill("6C3483"), font=_fnt(bold=True, size=12, color="FFFFFF"), aln=_CTR)
    for o, lbl in enumerate(["平均分", "平均%", "平均%圖表", "S.D."]):
        _sc(ws.cell(5, col+o), val=lbl,
            fill=_cfill("F0E6FA"), font=_fnt(bold=True, size=10, color="6C3483"), aln=_CTR)
    col += 4

    # vs 其他班（2欄）
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
    _sc(ws.cell(4, col), val=f"本班 vs {other_label}",
        fill=_cfill("6C3483"), font=_fnt(bold=True, size=12, color="FFFFFF"), aln=_CTR)
    for o, lbl in enumerate(["差距%", "差距圖表"]):
        _sc(ws.cell(5, col+o), val=lbl,
            fill=_cfill("E8DAEF"), font=_fnt(bold=True, size=10, color="6C3483"), aln=_CTR)

    # ── 資料行 ──
    prev_paper = None
    data_row = 6
    THIN = _thin_border()

    for ri, q in enumerate(q_cols):
        paper = paper_map[q]
        scale = scale_map[paper]
        mx    = float(max_scores[q])
        cs    = class_stats[cls][q]
        gs    = grade_stats[q]
        os_   = other_stats[q]
        bg    = _cfill("F7F9FB") if ri % 2 == 1 else NO_FILL

        # 卷別分隔行
        if paper != prev_paper:
            ws.row_dimensions[data_row].height = 15
            p_qs  = [q2 for q2 in q_cols if paper_map[q2] == paper]
            p_max = float(max_scores[p_qs].sum())
            sc_pct = int(round(scale * 100))
            ws.merge_cells(start_row=data_row, start_column=2,
                           end_row=data_row, end_column=1+NCOLS)
            _sc(ws.cell(data_row, 2),
                val=f"　{paper}　·　滿分 {int(p_max)}　·　差距圖表刻度 ±{sc_pct}%",
                fill=_cfill(dark_hex),
                font=_fnt(bold=True, size=11, color="FFFFFF"), aln=_LFT)
            data_row += 1
            prev_paper = paper

        ws.row_dimensions[data_row].height = 18
        dc = 2

        _sc(ws.cell(data_row, dc), val=q, fill=bg,
            font=_fnt(bold=True, size=11), aln=_CTR, bdr=THIN); dc += 1
        _sc(ws.cell(data_row, dc), val=int(mx), fill=bg,
            font=_fnt(size=11), aln=_CTR, bdr=THIN, fmt="0"); dc += 1

        # 本班
        _sc(ws.cell(data_row, dc), val=n_cls, fill=bg,
            font=_fnt(size=11), aln=_CTR, bdr=THIN, fmt="0"); dc += 1
        _sc(ws.cell(data_row, dc), val=cs["mean"], fill=bg,
            font=_fnt(size=11), aln=_CTR, bdr=THIN, fmt="0.00"); dc += 1
        _sc(ws.cell(data_row, dc), val=cs["mean_pct"], fill=bg,
            font=_fnt(size=11), aln=_CTR, bdr=THIN, fmt="0.0%"); dc += 1
        bar_c = ws.cell(data_row, dc)
        bar_c.value = _make_bar(cs["mean_pct"])
        bar_c.font  = Font(name="Courier New", size=10, color=bar_hex)
        bar_c.alignment = _LFT; bar_c.border = THIN; bar_c.fill = bg; dc += 1
        _sc(ws.cell(data_row, dc), val=cs["std"], fill=bg,
            font=_fnt(size=11), aln=_CTR, bdr=THIN, fmt="0.00"); dc += 1

        # 全級
        _sc(ws.cell(data_row, dc), val=gs["mean"], fill=_cfill("F2F3F4"),
            font=_fnt(size=11, color="3D4E5C"), aln=_CTR, bdr=THIN, fmt="0.00"); dc += 1
        _sc(ws.cell(data_row, dc), val=gs["mean_pct"], fill=_cfill("F2F3F4"),
            font=_fnt(size=11, color="3D4E5C"), aln=_CTR, bdr=THIN, fmt="0.0%"); dc += 1
        gb = ws.cell(data_row, dc)
        gb.value = _make_bar(gs["mean_pct"])
        gb.font  = Font(name="Courier New", size=10, color="7D8B99")
        gb.alignment = _LFT; gb.border = THIN; gb.fill = _cfill("F2F3F4"); dc += 1
        _sc(ws.cell(data_row, dc), val=gs["std"], fill=_cfill("F2F3F4"),
            font=_fnt(size=11, color="3D4E5C"), aln=_CTR, bdr=THIN, fmt="0.00"); dc += 1

        # vs 全級差距
        d1 = cs["mean_pct"] - gs["mean_pct"]
        _sc(ws.cell(data_row, dc), val=d1,
            fill=_diff_fill(d1), font=_fnt(bold=True, size=11, color=_diff_color(d1)),
            aln=_CTR, bdr=THIN, fmt="+0.0%;-0.0%;0.0%"); dc += 1
        BARS = 10
        filled = max(0, min(BARS, round(abs(d1) / scale * BARS)))
        if d1 >= 0:
            diff_bar = "▶" * filled + "·" * (BARS - filled)
            diff_col = "1A6B3A"
        else:
            diff_bar = "·" * (BARS - filled) + "◀" * filled
            diff_col = "9B1C1C"
        db1 = ws.cell(data_row, dc)
        db1.value = diff_bar
        db1.font  = Font(name="Courier New", size=11, color=diff_col)
        db1.alignment = _CTR; db1.border = THIN
        db1.fill  = _diff_fill(d1); dc += 1

        # 其他班
        _sc(ws.cell(data_row, dc), val=os_["mean"], fill=_cfill("F5EEF8"),
            font=_fnt(size=11, color="6C3483"), aln=_CTR, bdr=THIN, fmt="0.00"); dc += 1
        _sc(ws.cell(data_row, dc), val=os_["mean_pct"], fill=_cfill("F5EEF8"),
            font=_fnt(size=11, color="6C3483"), aln=_CTR, bdr=THIN, fmt="0.0%"); dc += 1
        ob = ws.cell(data_row, dc)
        ob.value = _make_bar(os_["mean_pct"])
        ob.font  = Font(name="Courier New", size=10, color="8E6BB0")
        ob.alignment = _LFT; ob.border = THIN; ob.fill = _cfill("F5EEF8"); dc += 1
        _sc(ws.cell(data_row, dc), val=os_["std"], fill=_cfill("F5EEF8"),
            font=_fnt(size=11, color="6C3483"), aln=_CTR, bdr=THIN, fmt="0.00"); dc += 1

        # vs 其他班差距
        d2 = cs["mean_pct"] - os_["mean_pct"]
        _sc(ws.cell(data_row, dc), val=d2,
            fill=_diff_fill(d2), font=_fnt(bold=True, size=11, color=_diff_color(d2)),
            aln=_CTR, bdr=THIN, fmt="+0.0%;-0.0%;0.0%"); dc += 1
        filled2 = max(0, min(BARS, round(abs(d2) / scale * BARS)))
        if d2 >= 0:
            diff_bar2 = "▶" * filled2 + "·" * (BARS - filled2)
            diff_col2 = "1A6B3A"
        else:
            diff_bar2 = "·" * (BARS - filled2) + "◀" * filled2
            diff_col2 = "9B1C1C"
        db2 = ws.cell(data_row, dc)
        db2.value = diff_bar2
        db2.font  = Font(name="Courier New", size=11, color=diff_col2)
        db2.alignment = _CTR; db2.border = THIN
        db2.fill  = _diff_fill(d2)

        data_row += 1

    # 欄寬
    widths = [8, 6,  7, 9, 9, 22, 8,  9, 9, 22, 8,  10, 14,  9, 9, 22, 8,  10, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(2+i)].width = w
    ws.freeze_panes = "B6"

# ══════════════════════════════════════════════════════════════
# 分數分佈工作表
# ══════════════════════════════════════════════════════════════
def _make_distribution_sheet(wb, df, max_scores, paper_map,
                              class_col, classes, palette):
    ws = wb.create_sheet("分數分佈")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    n_cls_cols = len(classes)
    total_cols = 2 + n_cls_cols * 2 + 2
    end_c = get_column_letter(1 + total_cols)

    ws.row_dimensions[2].height = 32
    ws.merge_cells(f"B2:{end_c}2")
    _sc(ws["B2"], val="分數分佈　·　各卷及合併總分",
        fill=_cfill("1F3864"),
        font=_fnt(bold=True, size=16, color="FFFFFF"), aln=_CTR)
    ws.row_dimensions[3].height = 14
    ws.merge_cells(f"B3:{end_c}3")
    _sc(ws["B3"],
        val="每 10% 為一分佈段（由高至低）。「高於本段人數」= 得分高於本段上限的人數。",
        font=_fnt(size=10, color="595959"), aln=_CTR, fill=NO_FILL)

    papers = sorted(set(paper_map.values()))
    q_cols = list(max_scores.index)

    sections = []
    for p in papers:
        p_qs   = [q for q in q_cols if paper_map[q] == p]
        p_max  = float(max_scores[p_qs].sum())
        p_raw  = df[p_qs].sum(axis=1)
        sections.append((p, p_raw, p_max))
    # 合併
    all_raw = df[q_cols].sum(axis=1)
    all_max = float(max_scores[q_cols].sum())
    if len(papers) > 1:
        sections.append(("合併", all_raw, all_max))

    THICK_B = _med_border("7F9AB5")
    STAT_B  = _thin_border("5D8AA8")

    cur_row = 4
    for sec_label, raw_series, s_max in sections:
        pass_th = s_max * 0.4

        # 段落標題
        ws.row_dimensions[cur_row].height = 22
        ws.merge_cells(start_row=cur_row, start_column=2,
                       end_row=cur_row, end_column=1+total_cols)
        _sc(ws.cell(cur_row, 2),
            val=f"◆  {sec_label}　　滿分：{int(s_max)}　　及格線：{round(pass_th)} 分（40%）",
            fill=_cfill("2C3E50"),
            font=_fnt(bold=True, size=12, color="FFFFFF"),
            aln=Alignment(horizontal="left", vertical="center", indent=1))
        cur_row += 1

        # 欄標頭
        ws.row_dimensions[cur_row].height = 20
        ws.row_dimensions[cur_row+1].height = 18
        col = 2
        for lbl in ["分佈（%）", "分數範圍"]:
            ws.merge_cells(start_row=cur_row, start_column=col,
                           end_row=cur_row+1, end_column=col)
            _sc(ws.cell(cur_row, col), val=lbl, fill=_cfill("D9D9D9"),
                font=_fnt(bold=True, size=11, color="1F3864"), aln=_CTR)
            col += 1
        for cls in classes:
            _, bg_hex, dark_hex = palette[cls]
            ws.merge_cells(start_row=cur_row, start_column=col,
                           end_row=cur_row, end_column=col+1)
            bar_hex = palette[cls][0]
            _sc(ws.cell(cur_row, col), val=cls,
                fill=_cfill(palette[cls][2]),
                font=_fnt(bold=True, size=11, color="FFFFFF"), aln=_CTR)
            for o, lbl in enumerate(["人數 (%)", "高於本段人數 (%)"]):
                _sc(ws.cell(cur_row+1, col+o), val=lbl,
                    fill=_cfill(bg_hex),
                    font=_fnt(bold=True, size=9, color=dark_hex), aln=_CTR)
            col += 2
        ws.merge_cells(start_row=cur_row, start_column=col,
                       end_row=cur_row, end_column=col+1)
        _sc(ws.cell(cur_row, col), val="全　級",
            fill=_cfill("5D6A77"),
            font=_fnt(bold=True, size=11, color="FFFFFF"), aln=_CTR)
        for o, lbl in enumerate(["人數 (%)", "高於本段人數 (%)"]):
            _sc(ws.cell(cur_row+1, col+o), val=lbl,
                fill=_cfill("EAECEE"),
                font=_fnt(bold=True, size=9, color="3D4E5C"), aln=_CTR)
        cur_row += 2

        # 分佈行（由高至低）
        bins = [i/10 for i in range(0, 11)]
        bin_labels = [f"{int(b*100)}%–{int((b+0.1)*100)-1}%" if b < 0.9 else "90%–100%"
                      for b in bins[:-1]]
        bin_lo = [round(b * s_max) for b in bins[:-1]]
        bin_hi = [round((b+0.1)*s_max)-1 for b in bins[:-1]]
        bin_hi[-1] = int(s_max)

        # 各班 raw series（同一 section）
        cls_raw = {}
        for cls in classes:
            cls_mask = class_col == cls
            if sec_label == "合併":
                cls_raw[cls] = df[cls_mask][q_cols].sum(axis=1)
            elif sec_label in set(paper_map.values()):
                p_qs = [q for q in q_cols if paper_map[q] == sec_label]
                cls_raw[cls] = df[cls_mask][p_qs].sum(axis=1)

        n_all = len(raw_series)
        for ri, bi in enumerate(range(len(bin_labels)-1, -1, -1)):
            lo, hi = bin_lo[bi], bin_hi[bi]
            row = cur_row + ri
            ws.row_dimensions[row].height = 17
            bg = _cfill("F7F9FB") if ri % 2 == 1 else NO_FILL
            col = 2
            _sc(ws.cell(row, col), val=bin_labels[bi], fill=_cfill("EBF5FB"),
                font=_fnt(bold=True, size=11, color="1A5276"), aln=_CTR); col += 1
            _sc(ws.cell(row, col), val=f"{lo}–{hi}", fill=_cfill("EBF5FB"),
                font=_fnt(size=11, color="1A5276"), aln=_CTR); col += 1
            for cls in classes:
                s = cls_raw[cls]
                n_c = len(s)
                cnt   = int(((s >= lo - 0.001) & (s <= hi + 0.001)).sum())
                above = int((s > hi + 0.001).sum())
                _, bg_hex, _ = palette[cls]
                _sc(ws.cell(row, col), val=f"{cnt} ({cnt/n_c:.0%})",
                    fill=bg, font=_fnt(size=11), aln=_CTR); col += 1
                _sc(ws.cell(row, col), val=f"{above} ({above/n_c:.0%})",
                    fill=bg, font=_fnt(size=11), aln=_CTR); col += 1
            cnt_a   = int(((raw_series >= lo-0.001) & (raw_series <= hi+0.001)).sum())
            above_a = int((raw_series > hi+0.001).sum())
            _sc(ws.cell(row, col), val=f"{cnt_a} ({cnt_a/n_all:.0%})",
                fill=_cfill("F2F3F4"),
                font=_fnt(bold=True, size=11, color="3D4E5C"), aln=_CTR); col += 1
            _sc(ws.cell(row, col), val=f"{above_a} ({above_a/n_all:.0%})",
                fill=_cfill("F2F3F4"),
                font=_fnt(bold=True, size=11, color="3D4E5C"), aln=_CTR)
        cur_row += len(bin_labels)

        # 合格人數/率
        ws.row_dimensions[cur_row].height = 20
        _set_merged_border(ws, cur_row, 2, cur_row, 3, "FEF9E7",
                           _fnt(bold=True, size=11, color="7D6608"),
                           f"合格人數 / 合格率（≥{round(pass_th)}分）", bdr=THICK_B)
        col = 4
        for cls in classes:
            s = cls_raw[cls]
            n_c = len(s)
            pss = int((s >= pass_th).sum())
            _set_merged_border(ws, cur_row, col, cur_row, col+1, "FEF9E7",
                               _fnt(bold=True, size=11, color="7D6608"),
                               f"{pss} ({pss/n_c:.1%})", bdr=THICK_B)
            col += 2
        pss_all = int((raw_series >= pass_th).sum())
        _set_merged_border(ws, cur_row, col, cur_row, col+1, "FEF9E7",
                           _fnt(bold=True, size=11, color="7D6608"),
                           f"{pss_all} ({pss_all/n_all:.1%})", bdr=THICK_B)
        cur_row += 1

        # 統計各一列
        stat_items = [
            ("平均分", lambda s: round(s.mean(), 1),     "0.0"),
            ("S.D.",   lambda s: round(s.std(ddof=1), 1), "0.0"),
            ("最高分", lambda s: int(s.max()),            "0"),
            ("最低分", lambda s: int(s.min()),            "0"),
        ]
        for stat_lbl, stat_fn, num_fmt in stat_items:
            ws.row_dimensions[cur_row].height = 17
            _set_merged_border(ws, cur_row, 2, cur_row, 3, "EBF5FB",
                               _fnt(bold=True, size=11, color="1A5276"),
                               stat_lbl, bdr=STAT_B)
            col = 4
            for cls in classes:
                s = cls_raw[cls]
                _, bg_hex, dark_hex = palette[cls]
                _set_merged_border(ws, cur_row, col, cur_row, col+1, bg_hex,
                                   _fnt(bold=True, size=11, color=dark_hex),
                                   stat_fn(s), fmt=num_fmt, bdr=STAT_B)
                col += 2
            _set_merged_border(ws, cur_row, col, cur_row, col+1, "EAECEE",
                               _fnt(bold=True, size=11, color="3D4E5C"),
                               stat_fn(raw_series), fmt=num_fmt, bdr=STAT_B)
            cur_row += 1

        cur_row += 2  # 段落間距

    # 欄寬
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    for ci in range(len(classes)):
        ws.column_dimensions[get_column_letter(4 + ci*2)].width = 14
        ws.column_dimensions[get_column_letter(5 + ci*2)].width = 16
    ws.column_dimensions[get_column_letter(4 + len(classes)*2)].width = 14
    ws.column_dimensions[get_column_letter(5 + len(classes)*2)].width = 16
    ws.freeze_panes = "B4"

# ══════════════════════════════════════════════════════════════
# 班際熱力圖工作表
# ══════════════════════════════════════════════════════════════
def _make_heatmap_sheet(wb, class_stats, grade_stats, classes, q_cols, paper_map, palette):
    ws = wb.create_sheet("班際熱力圖")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    NCOLS = 1 + len(classes) + 1
    end_c = get_column_letter(1 + NCOLS)

    ws.row_dimensions[2].height = 30
    ws.merge_cells(f"B2:{end_c}2")
    _sc(ws["B2"], val="班際熱力圖　·　各題平均得分率",
        fill=_cfill("1F3864"), font=_fnt(bold=True, size=15, color="FFFFFF"), aln=_CTR)

    ws.row_dimensions[3].height = 20
    _sc(ws["B3"], val="題號", fill=_cfill("D9D9D9"),
        font=_fnt(bold=True, size=11, color="1F3864"), aln=_CTR)
    for i, cls in enumerate(classes):
        _, _, dark_hex = palette[cls]
        _sc(ws.cell(3, 3+i), val=cls,
            fill=_cfill(palette[cls][2]),
            font=_fnt(bold=True, size=11, color="FFFFFF"), aln=_CTR)
    _sc(ws.cell(3, 3+len(classes)), val="全　級",
        fill=_cfill("5D6A77"), font=_fnt(bold=True, size=11, color="FFFFFF"), aln=_CTR)

    papers = sorted(set(paper_map.values()))
    prev_paper = None
    row = 4
    for q in q_cols:
        paper = paper_map[q]
        if paper != prev_paper:
            ws.row_dimensions[row].height = 14
            ws.merge_cells(start_row=row, start_column=2,
                           end_row=row, end_column=1+NCOLS)
            _, _, dark_hex_p = palette.get(classes[0], ("1F9E9E","E0F4F4","0D6E6E"))
            _sc(ws.cell(row, 2), val=f"  {paper}",
                fill=_cfill("2C3E50"),
                font=_fnt(bold=True, size=11, color="FFFFFF"), aln=_LFT)
            row += 1
            prev_paper = paper

        ws.row_dimensions[row].height = 17
        _sc(ws.cell(row, 2), val=q, fill=_cfill("EBF5FB"),
            font=_fnt(bold=True, size=11, color="1A5276"), aln=_CTR)

        pcts = [class_stats[cls][q]["mean_pct"] for cls in classes]
        pcts.append(grade_stats[q]["mean_pct"])
        min_p, max_p = min(pcts), max(pcts)

        for i, cls in enumerate(classes):
            pct = class_stats[cls][q]["mean_pct"]
            # 熱力色：低→紅，高→綠
            if max_p > min_p:
                t = (pct - min_p) / (max_p - min_p)
            else:
                t = 0.5
            r = int(255 - t * 150)
            g = int(180 + t * 75)
            b = int(150 - t * 50)
            hex_bg = f"{r:02X}{g:02X}{b:02X}"
            txt_color = "FFFFFF" if t < 0.35 or t > 0.85 else "1A1A1A"
            _sc(ws.cell(row, 3+i), val=pct,
                fill=_cfill(hex_bg),
                font=_fnt(size=11, color=txt_color), aln=_CTR, fmt="0.0%")

        g_pct = grade_stats[q]["mean_pct"]
        _sc(ws.cell(row, 3+len(classes)), val=g_pct,
            fill=_cfill("EAECEE"),
            font=_fnt(bold=True, size=11, color="3D4E5C"), aln=_CTR, fmt="0.0%")
        row += 1

    ws.column_dimensions["B"].width = 14
    for i in range(len(classes)+1):
        ws.column_dimensions[get_column_letter(3+i)].width = 12
    ws.freeze_panes = "B4"

# ══════════════════════════════════════════════════════════════
# 主要公開函數
# ══════════════════════════════════════════════════════════════
def generate_class_analysis_excel(
    df: pd.DataFrame,
    max_scores: pd.Series,
    paper_map: dict,
    class_info: pd.DataFrame,
    exam_info: dict,
) -> bytes:
    """
    生成全級班際分析 Excel（bytes）。

    參數
    ----
    df         : 成績 DataFrame（index=姓名，columns=題目欄名）
    max_scores : 各題滿分 Series
    paper_map  : {題目欄名: "P1"/"P2"/...}
    class_info : DataFrame，含 "班別" 欄（index 需可對應 df.index 或另有 "中文姓名" 欄）
    exam_info  : dict，至少含 "exam_title"（用於標題）
    """
    # ── 建立班別 Series（index 對齊 df）──
    if "中文姓名" in class_info.columns:
        cls_map = dict(zip(class_info["中文姓名"].astype(str),
                           class_info["班別"].astype(str)))
        class_col = pd.Series(
            [cls_map.get(str(n), "未知") for n in df.index],
            index=df.index, name="班別"
        )
    else:
        class_col = class_info["班別"].astype(str)
        class_col.index = df.index

    q_cols  = list(max_scores.index)
    classes = sorted(class_col.unique().tolist())
    palette = _build_palette(classes)

    class_stats, grade_stats, classes, paper_totals, combined_totals = _compute_stats(
        df, q_cols, max_scores, paper_map, class_col
    )

    wb = Workbook()
    wb.remove(wb.active)

    # 各班分析工作表
    for cls in classes:
        _make_class_sheet(wb, cls, q_cols, max_scores, paper_map,
                          class_stats, grade_stats, df, class_col, palette)

    # 分數分佈工作表
    _make_distribution_sheet(wb, df, max_scores, paper_map,
                             class_col, classes, palette)

    # 班際熱力圖工作表
    _make_heatmap_sheet(wb, class_stats, grade_stats, classes,
                        q_cols, paper_map, palette)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def get_class_summary_df(
    df: pd.DataFrame,
    max_scores: pd.Series,
    paper_map: dict,
    class_info: pd.DataFrame,
) -> pd.DataFrame:
    """
    回傳班際摘要 DataFrame（供 Streamlit 網頁預覽）。
    欄位：班別 | 人數 | 合併平均分 | 合併平均% | 合格率
    """
    if "中文姓名" in class_info.columns:
        cls_map = dict(zip(class_info["中文姓名"].astype(str),
                           class_info["班別"].astype(str)))
        class_col = pd.Series(
            [cls_map.get(str(n), "未知") for n in df.index],
            index=df.index
        )
    else:
        class_col = class_info["班別"].astype(str)

    q_cols   = list(max_scores.index)
    total_mx = float(max_scores.sum())
    pass_th  = total_mx * 0.4
    rows = []
    for cls in sorted(class_col.unique()):
        sub = df[class_col == cls][q_cols]
        raw = sub.sum(axis=1)
        rows.append({
            "班別":     cls,
            "人數":     len(sub),
            "合併平均分": round(raw.mean(), 1),
            "合併平均%":  f"{raw.mean()/total_mx*100:.1f}%",
            "合格率":    f"{(raw >= pass_th).mean()*100:.1f}%",
        })
    # 全級
    all_raw = df[q_cols].sum(axis=1)
    rows.append({
        "班別":     "全　級",
        "人數":     len(df),
        "合併平均分": round(all_raw.mean(), 1),
        "合併平均%":  f"{all_raw.mean()/total_mx*100:.1f}%",
        "合格率":    f"{(all_raw >= pass_th).mean()*100:.1f}%",
    })
    return pd.DataFrame(rows)
