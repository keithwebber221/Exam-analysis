# exam_item_analysis.py
# 依賴安裝：pip install pandas openpyxl plotly kaleido scipy python-docx pillow
# 執行方式：python exam_item_analysis.py

import os
import sys
import io
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from scipy import stats as scipy_stats

# ── matplotlib 中文字型設定 ──
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib import font_manager, rcParams

def _setup_cjk_font():
    """動態偵測中文字型（Linux Streamlit Cloud / macOS / Windows）"""
    candidates = [
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSerifCJK-Regular.ttc",
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/noto-cjk/NotoSansCJKtc-Regular.otf",
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Light.ttc",
        "/Library/Fonts/Arial Unicode.ttf",
        "C:/Windows/Fonts/msjh.ttc",
        "C:/Windows/Fonts/mingliu.ttc",
        "C:/Windows/Fonts/simsun.ttc",
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                font_manager.fontManager.addfont(path)
                prop = font_manager.FontProperties(fname=path)
                name = prop.get_name()
                rcParams["font.family"]     = "sans-serif"
                rcParams["font.sans-serif"] = [name] + rcParams.get("font.sans-serif", [])
                rcParams["axes.unicode_minus"] = False
                return name
            except Exception:
                continue
    rcParams["axes.unicode_minus"] = False
    return "DejaVu Sans"

_CJK_FONT = _setup_cjk_font()

TOP_BOTTOM_PCT = 0.27


# ============================================================
# 0. 互動式啟動設定
# ============================================================
def get_exam_info() -> dict:
    print("=" * 55)
    print("   📋 DSE 試卷分析系統")
    print("=" * 55)

    while True:
        year_input = input("\n📅 請輸入試卷年份（例：2025-2026 年度請填 2526）：").strip()
        if len(year_input) == 4 and year_input.isdigit():
            first_two = int(year_input[:2])
            last_two  = int(year_input[2:])
            if last_two == first_two + 1:
                year_label = f"20{year_input[:2]}-20{year_input[2:]}"
                break
            else:
                print("   ⚠️  格式錯誤！例如 2025-2026 年度請填「2526」，請重新輸入。")
        else:
            print("   ⚠️  請輸入 4 位數字，例如「2526」代表 2025-2026 年度。")

    EXAM_TYPES = {"1": "上學期測驗", "2": "上學期考試", "3": "下學期測驗", "4": "下學期考試"}
    EXAM_CODES = {"1": "T1T", "2": "T1E", "3": "T2T", "4": "T2E"}

    print("\n📂 請選擇試卷類別：")
    for key, label in EXAM_TYPES.items():
        print(f"   [{key}] {label}")

    while True:
        type_input = input("\n請輸入選項編號（1-4）：").strip()
        if type_input in EXAM_TYPES:
            exam_type_label = EXAM_TYPES[type_input]
            exam_type_code  = EXAM_CODES[type_input]
            break
        else:
            print("   ⚠️  請輸入 1、2、3 或 4。")

    subject_input = input("\n📚 請輸入科目名稱（選填，按 Enter 跳過）：").strip()
    subject_label = subject_input if subject_input else ""

    FORM_LEVELS = {"1": "F1", "2": "F2", "3": "F3", "4": "F4", "5": "F5", "6": "F6"}
    print("\n🎓 請選擇年級：")
    for key, label in FORM_LEVELS.items():
        print(f"   [{key}] {label}", end="   " if int(key) % 3 != 0 else "\n")
    print()
    while True:
        form_input = input("\n請輸入選項編號（1-6）：").strip()
        if form_input in FORM_LEVELS:
            form_label = FORM_LEVELS[form_input]
            break
        else:
            print("   ⚠️  請輸入 1 至 6。")

    print("\n📊 請選擇及格分數線：")
    print("   [1] 40%（適用：高中科目）")
    print("   [2] 50%（適用：初中科目）")
    while True:
        pass_input = input("\n請輸入選項編號（1 或 2，預設 1）：").strip()
        if pass_input == "" or pass_input == "1":
            pass_rate = 0.4
            pass_label = "40%"
            break
        elif pass_input == "2":
            pass_rate = 0.5
            pass_label = "50%"
            break
        else:
            print("   ⚠️  請輸入 1 或 2。")

    # 組合標題和檔名
    subject_display = subject_label if subject_label else "未指定科目"
    exam_title  = f"{year_label} {exam_type_label}｜{form_label} {subject_display}"

    # 檔名格式：2526_T2T_F5_BAFS（科目空白時省略）
    subject_slug = f"_{subject_label}" if subject_label else ""
    file_prefix = f"{year_input}_{exam_type_code}_{form_label}{subject_slug}"

    print("\n" + "─" * 55)
    print("✅ 確認考試資訊：")
    print(f"   年度：{year_label}")
    print(f"   類別：{exam_type_label}")
    print(f"   年級：{form_label}")
    print(f"   科目：{subject_display}")
    print(f"   及格線：{pass_label}")
    print(f"   輸出檔案前綴：{file_prefix}")
    print("─" * 55)

    # ── 試卷數目與比例 ──
    print("\n📄 請設定試卷結構：")
    print("   [1] 單一試卷（無需分卷）")
    print("   [2] 兩份試卷（卷一 + 卷二）")
    print("   [3] 三份試卷（卷一 + 卷二 + 卷三）")
    print("   [4] 四份試卷（卷一 + 卷二 + 卷三 + 卷四）")
    while True:
        paper_input = input("\n請輸入選項（1/2/3/4，預設 1）：").strip()
        if paper_input in ["", "1", "2", "3", "4"]:
            num_papers = 1 if paper_input in ["", "1"] else int(paper_input)
            break
        print("   ⚠️  請輸入 1、2、3 或 4。")

    paper_weights = {}
    if num_papers == 1:
        paper_weights = {"P1": 1.0}
        print("   ✅ 單一試卷，無加權設定")
    else:
        print(f"\n   請輸入各試卷佔分比例（合計必須為 100%）：")
        while True:
            weights_ok = True
            total_w = 0
            tmp = {}
            for p in range(1, num_papers + 1):
                while True:
                    w = input(f"   卷{p}（P{p}）佔分比例（整數 %）：").strip()
                    if w.isdigit() and 1 <= int(w) <= 100:
                        tmp[f"P{p}"] = int(w)
                        total_w += int(w)
                        break
                    print("   ⚠️  請輸入 1-100 的整數。")
            if total_w == 100:
                paper_weights = {k: v / 100 for k, v in tmp.items()}
                for p, w in tmp.items():
                    print(f"   ✅ {p}：{w}%")
                break
            else:
                print(f"   ⚠️  各卷比例合計為 {total_w}%，必須等於 100%，請重新輸入。")

    confirm = input("\n確認無誤？按 Enter 繼續，輸入 N 重新設定：").strip().upper()
    if confirm == "N":
        return get_exam_info()

    return {
        "year_label": year_label, "exam_type_label": exam_type_label,
        "exam_type_code": exam_type_code, "subject_label": subject_label,
        "form_label": form_label, "exam_title": exam_title,
        "file_prefix": file_prefix, "pass_rate": pass_rate, "pass_label": pass_label,
        "paper_weights": paper_weights, "num_papers": num_papers,
    }


# ============================================================
# 1. 讀取資料
# ============================================================
def load_data(filepath: str):
    """
    讀取 scores.xlsx，自動偵測新舊格式：
    新格式（多試卷）：第4行為【試卷】行（P1/P2/...），第5行為【滿分】，第6行起為學生
    舊格式（單試卷）：第4行為【滿分】，第5行起為學生
    回傳：(score_data, max_scores, absent_set, paper_map)
      paper_map: dict {題目欄名 -> "P1"/"P2"/...}，單試卷全部為 "P1"
    """
    df_raw    = pd.read_excel(filepath, header=None)
    col_names = df_raw.iloc[2].tolist()

    # ── 偵測第4行是否為【試卷】行 ──
    row3_vals = df_raw.iloc[3].tolist()
    # 第4行的非空值中，若含有 P1/P2/P3/P4 等，視為試卷行
    paper_labels = {"P1", "P2", "P3", "P4"}
    row3_str = [str(v).strip() for v in row3_vals if pd.notna(v) and str(v).strip() != ""]
    has_paper_row = any(v in paper_labels for v in row3_str)

    if has_paper_row:
        paper_row   = row3_vals          # 試卷行
        max_row     = df_raw.iloc[4].tolist()  # 滿分行
        student_raw = df_raw.iloc[5:].copy()
        print("   📋 偵測到【試卷】行，啟用多試卷模式")
    else:
        paper_row   = None               # 無試卷行（單試卷）
        max_row     = row3_vals          # 滿分行
        student_raw = df_raw.iloc[4:].copy()

    student_raw.columns = col_names
    student_raw = student_raw[
        student_raw["中文姓名"].notna() &
        ~student_raw["中文姓名"].astype(str).str.contains("說明|輸入", na=False)
    ]

    # ── 識別缺席學生 ──
    absent_set = set()
    absent_col_candidates = ["缺席", "Absent", "absent", "缺考"]
    absent_col = next((c for c in col_names if str(c) in absent_col_candidates), None)
    if absent_col:
        for _, row in student_raw.iterrows():
            val = row.get(absent_col, "")
            if pd.notna(val) and str(val).strip() not in ["", "0", "nan"]:
                absent_set.add(str(row["中文姓名"]).strip())
        if absent_set:
            print(f"   ✋ 缺席學生（{len(absent_set)} 人）：{'、'.join(sorted(absent_set))}")

    info_cols     = ["班別", "班號", "英文姓名", "中文姓名"] + absent_col_candidates
    question_cols = [c for c in col_names
                     if isinstance(c, str) and c.strip() != "" and c not in info_cols]
    max_scores = pd.Series(max_row, index=col_names)[question_cols].astype(float)
    score_data = student_raw[question_cols].astype(float)
    score_data.index = student_raw["中文姓名"].values
    score_data.index.name = "姓名"
    score_data.fillna(0, inplace=True)

    # ── 建立 paper_map：題目 → 試卷編號 ──
    paper_map = {}
    if has_paper_row:
        paper_series = pd.Series(paper_row, index=col_names)
        for q in question_cols:
            val = str(paper_series.get(q, "P1")).strip()
            paper_map[q] = val if val in paper_labels else "P1"
        papers_found = sorted(set(paper_map.values()))
        for p in papers_found:
            cnt = sum(1 for v in paper_map.values() if v == p)
            print(f"   · {p}：{cnt} 題")
    else:
        paper_map = {q: "P1" for q in question_cols}

    return score_data, max_scores, absent_set, paper_map
# ============================================================
# 1b. 多試卷工具函數
# ============================================================
def get_paper_groups(columns, paper_weights, paper_map=None):
    """
    依 paper_map（試卷行讀取）或 paper_weights 鍵分組題目。
    paper_map: {題目欄名 -> "P1"/"P2"/...}（load_data 回傳）
    若無 paper_map 或只有單試卷，全部歸 P1。
    回傳 dict: {"P1": [col,...], "P2": [...], ...}
    """
    groups = {p: [] for p in paper_weights}
    default_p = list(paper_weights.keys())[0]

    if paper_map and len(paper_weights) > 1:
        for col in columns:
            p = paper_map.get(col, default_p)
            if p in groups:
                groups[p].append(col)
            else:
                groups[default_p].append(col)
    else:
        groups[default_p] = list(columns)
    return groups


def calc_weighted_scores(df, max_scores, paper_weights, paper_map=None):
    """
    計算加權總分（滿分統一為 100 分制）。
    每卷得分率 × 該卷權重 × 100，合計為加權總分。
    同時回傳各卷得分率 DataFrame（學生 × 試卷）。
    """
    groups = get_paper_groups(df.columns.tolist(), paper_weights, paper_map)
    paper_pct = pd.DataFrame(index=df.index)
    weighted = pd.Series(0.0, index=df.index)
    for p, cols in groups.items():
        if not cols:
            paper_pct[p] = np.nan
            continue
        p_max = max_scores[cols].sum()
        p_raw = df[cols].sum(axis=1)
        pct   = (p_raw / p_max * 100).round(2) if p_max > 0 else pd.Series(0.0, index=df.index)
        paper_pct[p]  = pct
        weighted      += pct * paper_weights[p]
    weighted = weighted.round(2)
    total_max_weighted = 100.0   # 加權後滿分固定為 100
    return weighted, paper_pct, total_max_weighted




# ============================================================
# 2. 試題分析
# ============================================================
def classify_difficulty(P): 
    return "🟢 容易" if P >= 0.75 else ("🟡 適中" if P >= 0.25 else "🔴 困難")

def classify_discrimination(D):
    return "⭐ 優良" if D >= 0.40 else ("✅ 良好" if D >= 0.30 else ("⚠️ 尚可" if D >= 0.20 else "❌ 不佳"))

def suggest_action(P, D):
    if P < 0.25 and D < 0.20: return "建議修改題目（太難且鑑別差）"
    if P > 0.75 and D < 0.20: return "建議替換題目（太易且鑑別差）"
    if D < 0.20:               return "建議檢視題目表述"
    if P < 0.25:               return "全班需加強此課題"
    return "題目質素良好"

def item_analysis(df: pd.DataFrame, max_scores: pd.Series,
                  absent_set: set = None) -> pd.DataFrame:
    # 排除缺席學生後再計算 P 值和 D 值
    df_active = df[~df.index.isin(absent_set or set())]
    n = len(df_active)
    k = max(1, int(np.floor(n * TOP_BOTTOM_PCT)))
    df_temp = df_active.copy()
    df_temp["total"] = df_temp.sum(axis=1)
    sorted_df  = df_temp.sort_values("total", ascending=False)
    high_group = sorted_df.head(k).drop(columns="total")
    low_group  = sorted_df.tail(k).drop(columns="total")
    results = []
    skipped = []
    for q in df_active.columns:
        max_s = max_scores.get(q, np.nan) if hasattr(max_scores, 'get') else max_scores[q]
        # 跳過滿分為 NaN 或 0 的題目
        if pd.isna(max_s) or max_s <= 0:
            skipped.append(q)
            continue
        scores    = df_active[q]
        P         = scores.mean() / max_s
        threshold = max_s * 0.5
        PH = (high_group[q] >= threshold).mean()
        PL = (low_group[q]  >= threshold).mean()
        D  = round(PH - PL, 3)
        results.append({
            "題號": q, "滿分": int(max_s), "平均分": round(scores.mean(), 2),
            "標準差": round(scores.std(), 2), "最高分": scores.max(), "最低分": scores.min(),
            "難度指數 P": round(P, 3), "鑑別度 D": D,
            "難度評級": classify_difficulty(P), "鑑別評級": classify_discrimination(D),
            "建議行動": suggest_action(P, D),
        })
    if skipped:
        print(f"   ⚠️  以下題目的滿分為空白或零，已略過分析：{', '.join(str(q) for q in skipped)}")
        print(f"      請確認 scores.xlsx【滿分】行中對應欄位已填寫數值")
    return pd.DataFrame(results)


# ============================================================
# 3. 學生總分分析
# ============================================================
def student_summary(df: pd.DataFrame, max_scores: pd.Series,
                    pass_rate: float = 0.4, absent_set: set = None,
                    paper_weights: dict = None, paper_pct: pd.DataFrame = None,
                    weighted_scores: pd.Series = None, num_papers: int = 1):
    absent_set  = absent_set or set()
    df2         = df.copy()

    # 單卷：原始總分；多卷：加權後分數（0-100制）
    if num_papers > 1 and paper_weights and weighted_scores is not None:
        total_max   = 100.0
        df2["總分"]    = weighted_scores
        df2["百分比(%)"] = weighted_scores.round(1)
    else:
        total_max   = float(max_scores.sum())
        df2["總分"]    = df2[max_scores.index].sum(axis=1)
        df2["百分比(%)"] = (df2["總分"] / total_max * 100).round(1)

    # 排名只計算出席學生
    score_col_name = "總分(加權)" if ("總分(加權)" in df2.columns) else "總分"
    df_active      = df2[~df2.index.isin(absent_set)]
    df2["排名"]     = df2[score_col_name].rank(ascending=False, method="min").astype(int)

    # 缺席欄：出席/缺席
    df2["出席狀態"] = df2.index.map(lambda n: "缺席" if n in absent_set else "出席")

    # 加入各試卷得分率欄（多試卷才顯示）
    if num_papers > 1 and paper_weights and paper_pct is not None:
        for p in paper_weights:
            if p in paper_pct.columns:
                pct_col = f"{p}得分率(%)"
                df2[pct_col] = paper_pct[p].reindex(df2.index).round(1)
        paper_cols = [f"{p}得分率(%)" for p in paper_weights if f"{p}得分率(%)" in df2.columns]
        summary_cols = ["出席狀態"] + paper_cols + ["總分(加權)", "百分比(%)", "排名"]
        df2.rename(columns={"總分": "總分(加權)"}, inplace=True)
    else:
        summary_cols = ["出席狀態", "總分", "百分比(%)", "排名"]
    summary = df2[summary_cols].copy()
    # 轉為 object 型別，允許混合數字和文字（缺席標示）
    # 轉 object 型別，允許混合數字和文字（缺席標示）
    for _col in ["總分", "總分(加權)", "百分比(%)", "排名"]:
        if _col in summary.columns:
            summary[_col] = summary[_col].astype(object)
    # 缺席學生的總分/排名欄顯示「缺席」
    for name in absent_set:
        if name in summary.index:
            for _c in ["總分", "總分(加權)"]:
                if _c in summary.columns:
                    summary.loc[name, _c] = "缺席"
            summary.loc[name, "百分比(%)"] = "缺席"
            summary.loc[name, "排名"]      = "-"
            # 各卷得分率亦標示缺席
            for _c in summary.columns:
                if "得分率" in _c:
                    summary.loc[name, _c] = "缺席"
    summary = summary.sort_values("出席狀態", key=lambda x: x.map({"出席": 0, "缺席": 1}))

    # 全班統計只計算出席學生
    n_total   = len(df2)
    n_absent  = len(absent_set & set(df2.index))
    n_present = n_total - n_absent

    stats = pd.DataFrame({
        "統計項目": [
            "全班人數", "出席人數", "缺席人數", "滿分",
            "平均分（出席）", "中位數（出席）", "標準差（出席）",
            "最高分", "最低分",
            f"及格率（≥{int(pass_rate*100)}%，出席）"
        ],
        "數值": [
            n_total, n_present, n_absent, int(total_max),
            round(df_active[score_col_name].mean(), 1)  if n_present > 0 else "-",
            df_active[score_col_name].median()           if n_present > 0 else "-",
            round(df_active[score_col_name].std(), 2)    if n_present > 0 else "-",
            df_active[score_col_name].max()              if n_present > 0 else "-",
            df_active[score_col_name].min()              if n_present > 0 else "-",
            f"{(df_active[score_col_name] >= total_max * pass_rate).mean()*100:.1f}%"
            if n_present > 0 else "-"
        ]
    })
    return summary, stats


# ============================================================
# 4. 大題分析
# ============================================================
def question_group_analysis(df, max_scores, item_df):
    def get_parent(q):
        q = str(q)   # 防禦：確保是字串，避免 numpy.float64 等型別報錯
        return q.rstrip("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ")
    groups = {}
    for q in df.columns:
        if not isinstance(q, str) or q.strip() == "":
            continue  # 跳過非字串或空白欄名
        groups.setdefault(get_parent(q), []).append(q)
    rows = []
    for parent, sub_qs in groups.items():
        gmax = max_scores[sub_qs].sum()
        gavg = df[sub_qs].sum(axis=1).mean()
        avgP = round(item_df[item_df["題號"].isin(sub_qs)]["難度指數 P"].mean(), 3)
        rows.append({
            "大題": parent, "分題數": len(sub_qs), "大題滿分": int(gmax),
            "全班平均": round(gavg, 2), "平均難度 P": avgP,
            "得分率": f"{gavg/gmax*100:.1f}%" if gmax > 0 else "N/A",
            "難度評級": classify_difficulty(avgP),
        })
    return pd.DataFrame(rows)


# ============================================================
# 5. 圖表
# ============================================================
def create_charts(df, max_scores, item_df, student_df, exam_title,
                  chart_dir=None, absent_set=None, return_bytes=False):
    """
    用 matplotlib 生成4張分析圖表（無需 kaleido / Chrome）。
    chart_dir: 儲存到資料夾（None 則不寫檔）
    return_bytes: True 則回傳 dict {檔名: PNG bytes}
    """
    import matplotlib.colors as mcolors
    from matplotlib.patches import FancyBboxPatch
    import matplotlib.ticker as mticker

    absent_set = absent_set or set()
    df_plot    = df[~df.index.isin(absent_set)]
    item_df    = item_df.copy()
    item_df["得分率 %"] = (item_df["平均分"] / item_df["滿分"] * 100).round(1)
    charts_bytes = {}

    # 圖表用純文字（避免 matplotlib 無法渲染 emoji）
    COLOR     = {"容易": "#2ecc71", "適中": "#f39c12", "困難": "#e74c3c"}
    # emoji → 純文字 對照（item_df 的「難度評級」欄含 emoji，需轉換）
    EMOJI_MAP = {"🟢 容易": "容易", "🟡 適中": "適中", "🔴 困難": "困難"}
    item_df["難度評級_圖"] = item_df["難度評級"].map(EMOJI_MAP).fillna(item_df["難度評級"])

    def _fig_to_bytes(fig):
        buf = io.BytesIO()
        fig.savefig(buf, format="png", dpi=180, bbox_inches="tight",
                    facecolor=fig.get_facecolor())
        plt.close(fig)
        return buf.getvalue()

    def _save(fname, fig):
        img = _fig_to_bytes(fig)
        charts_bytes[fname] = img
        if chart_dir:
            os.makedirs(chart_dir, exist_ok=True)
            with open(f"{chart_dir}/{fname}", "wb") as fh:
                fh.write(img)

    # ── 圖1：難度－鑑別度散點圖 ──
    fig1, ax1 = plt.subplots(figsize=(11, 7))
    fig1.patch.set_facecolor("#FAFAFA")
    ax1.set_facecolor("#F5F7FA")
    for label, color in COLOR.items():
        sub = item_df[item_df["難度評級_圖"] == label]
        if len(sub):
            ax1.scatter(sub["難度指數 P"], sub["鑑別度 D"],
                        c=color, label=label, s=120, zorder=4,
                        edgecolors="white", linewidths=0.8)
            for _, r in sub.iterrows():
                ax1.annotate(str(r["題號"]),
                             (r["難度指數 P"], r["鑑別度 D"]),
                             textcoords="offset points", xytext=(7, 4),
                             fontsize=8, zorder=5)
    ax1.axhline(0.4, ls="--", c="#27ae60", lw=1.2, label="D=0.4（優良）")
    ax1.axhline(0.2, ls="--", c="#e67e22", lw=1.2, label="D=0.2（尚可）")
    ax1.axvline(0.25, ls=":",  c="#e74c3c", lw=1.2, label="P=0.25（困難）")
    ax1.axvline(0.75, ls=":",  c="#3498db", lw=1.2, label="P=0.75（容易）")
    ax1.set_xlabel("難度指數 P", fontsize=12)
    ax1.set_ylabel("鑑別度 D",   fontsize=12)
    ax1.set_title(f"{exam_title}｜題目難度 vs 鑑別度", fontsize=14, fontweight="bold", pad=12)
    ax1.legend(fontsize=9, loc="upper left", framealpha=0.8)
    ax1.grid(True, alpha=0.3)
    ax1.set_xlim(-0.05, 1.05); ax1.set_ylim(-0.15, 1.05)
    _save("01_difficulty_discrimination.png", fig1)

    # ── 圖2：各題得分率橫條圖 ──
    item_sorted = item_df.sort_values("得分率 %")
    bar_colors  = [COLOR.get(EMOJI_MAP.get(d, d), "#95a5a6") for d in item_sorted["難度評級"]]
    fig2, ax2   = plt.subplots(figsize=(11, max(5, len(item_sorted) * 0.38)))
    fig2.patch.set_facecolor("#FAFAFA")
    ax2.set_facecolor("#F5F7FA")
    bars = ax2.barh(item_sorted["題號"].astype(str),
                    item_sorted["得分率 %"],
                    color=bar_colors, edgecolor="white", linewidth=0.6, height=0.7)
    for bar, val in zip(bars, item_sorted["得分率 %"]):
        ax2.text(val + 0.8, bar.get_y() + bar.get_height() / 2,
                 f"{val:.1f}%", va="center", fontsize=8, color="#333")
    ax2.axvline(50, ls="--", c="gray", lw=1.2, label="50% 基準", alpha=0.8)
    ax2.set_xlabel("得分率 %", fontsize=12)
    ax2.set_title(f"{exam_title}｜各題得分率排行", fontsize=14, fontweight="bold", pad=12)
    ax2.set_xlim(0, min(130, item_sorted["得分率 %"].max() + 18))
    ax2.legend(fontsize=9)
    ax2.grid(axis="x", alpha=0.3)
    # 圖例：難度顏色
    from matplotlib.patches import Patch
    legend_els = [Patch(facecolor=c, label=f"◼ {l}") for l, c in COLOR.items()]
    ax2.legend(handles=legend_els + [
        plt.Line2D([0],[0], ls="--", c="gray", label="50% 基準")
    ], fontsize=9, loc="lower right")
    _save("02_score_rate_by_question.png", fig2)

    # ── 圖3：全班總分分佈 + KDE 曲線 ──
    _score_col    = "總分(加權)" if "總分(加權)" in student_df.columns else "總分"
    scores_num    = pd.to_numeric(student_df[_score_col], errors="coerce").dropna()
    scores_arr    = scores_num.values.astype(float)
    total_max_val = 100 if _score_col == "總分(加權)" else int(max_scores.sum())
    fig3, ax3     = plt.subplots(figsize=(10, 5.5))
    fig3.patch.set_facecolor("#FAFAFA")
    ax3.set_facecolor("#F5F7FA")
    if len(scores_arr) >= 2:
        n, bins, patches = ax3.hist(scores_arr, bins=10,
                                     color="#3498db", alpha=0.6,
                                     edgecolor="white", linewidth=0.8,
                                     label="人數分佈")
        if scores_arr.max() > scores_arr.min():
            x_range = np.linspace(scores_arr.min(), scores_arr.max(), 300)
            kde      = scipy_stats.gaussian_kde(scores_arr)
            density  = kde(x_range) * len(scores_arr) * (scores_arr.max()-scores_arr.min()) / 10
            ax3.plot(x_range, density, color="#e74c3c", lw=2.5, label="密度曲線")
        mean_v = scores_arr.mean()
        ax3.axvline(mean_v, color="#27ae60", ls="--", lw=2,
                    label=f"平均 {mean_v:.1f}")
    ax3.set_xlabel(f"總分（滿分 {total_max_val}）", fontsize=12)
    ax3.set_ylabel("人數", fontsize=12)
    ax3.set_title(f"{exam_title}｜全班總分分佈", fontsize=14, fontweight="bold", pad=12)
    ax3.legend(fontsize=10); ax3.grid(alpha=0.3)
    _save("03_student_score_distribution.png", fig3)

    # ── 圖4：全班答題熱力圖 ──
    score_pct = df_plot.div(max_scores) * 100
    n_students, n_q = score_pct.shape
    fig_w = max(12, n_q * 0.55)
    fig_h = max(6,  n_students * 0.32)
    fig4, ax4 = plt.subplots(figsize=(fig_w, fig_h))
    fig4.patch.set_facecolor("#FAFAFA")

    cmap = mcolors.LinearSegmentedColormap.from_list(
        "RdYlGn", ["#e74c3c","#f39c12","#f1c40f","#2ecc71"], N=256)
    im = ax4.imshow(score_pct.values, aspect="auto",
                    cmap=cmap, vmin=0, vmax=100, interpolation="nearest")
    cbar = fig4.colorbar(im, ax=ax4, shrink=0.8, pad=0.02)
    cbar.set_label("得分率 %", fontsize=10)

    # 數值標注（學生數≤40 才顯示，避免擁擠）
    if n_students <= 40 and n_q <= 50:
        fs = max(5, min(8, 200 // (n_students * n_q + 1)))
        for i in range(n_students):
            for j in range(n_q):
                val = score_pct.values[i, j]
                if not np.isnan(val):
                    txt_color = "white" if val < 40 or val > 80 else "black"
                    ax4.text(j, i, f"{val:.0f}", ha="center", va="center",
                             fontsize=fs, color=txt_color)

    ax4.set_xticks(range(n_q))
    ax4.set_xticklabels(score_pct.columns.astype(str),
                        rotation=-45, ha="right", fontsize=max(7, min(9, 220//n_q)))
    ax4.set_yticks(range(n_students))
    ax4.set_yticklabels(score_pct.index.astype(str),
                        fontsize=max(6, min(9, 180//n_students)))
    ax4.set_xlabel("試題", fontsize=12)
    ax4.set_ylabel("學生", fontsize=12)
    ax4.set_title(f"{exam_title}｜全班答題熱力圖（得分率 %）",
                  fontsize=14, fontweight="bold", pad=12)
    plt.tight_layout()
    _save("04_class_heatmap.png", fig4)

    if chart_dir:
        print(f"   ✅ 4 張圖表已儲存至 {chart_dir}/")
    if return_bytes:
        return charts_bytes


# ============================================================
# 6. 匯出 Excel
# ============================================================
def export_excel(item_df, group_df, student_df, stats_df, output_path, exam_title):
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"),  bottom=Side(style="thin"))
    header_fill = PatternFill("solid", fgColor="1F4788")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    color_fills = {
        "🟢 容易": PatternFill("solid", fgColor="C6EFCE"),
        "🟡 適中": PatternFill("solid", fgColor="FFEB9C"),
        "🔴 困難": PatternFill("solid", fgColor="FFC7CE"),
        "⭐ 優良": PatternFill("solid", fgColor="C6EFCE"),
        "✅ 良好": PatternFill("solid", fgColor="DDEBF7"),
        "⚠️ 尚可": PatternFill("solid", fgColor="FFEB9C"),
        "❌ 不佳": PatternFill("solid", fgColor="FFC7CE"),
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        item_df.to_excel(writer,    sheet_name="1_試題分析", index=False)
        group_df.to_excel(writer,   sheet_name="2_大題分析", index=False)

        # 學生成績：出席學生按排名升序，缺席學生排在最後
        student_df_sorted = student_df.copy()
        rank_col = pd.to_numeric(student_df_sorted["排名"], errors="coerce")
        student_df_sorted = student_df_sorted.iloc[rank_col.argsort(kind="stable")]
        absent_mask = student_df_sorted["出席狀態"] == "缺席"
        student_df_sorted = pd.concat([
            student_df_sorted[~absent_mask],
            student_df_sorted[absent_mask]
        ])
        # 若有加權總分，在表頭加說明
        score_col = "總分(加權)" if "總分(加權)" in student_df_sorted.columns else "總分"
        student_df_sorted.to_excel(writer, sheet_name="3_學生成績", index=True)

        stats_df.to_excel(writer,   sheet_name="4_全班統計", index=False)

        wb = writer.book

        # 說明工作表（最後）
        ws_leg = wb.create_sheet("5_📖 說明")
        ws_leg["A1"] = "試題分析指標說明"
        ws_leg["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws_leg["A1"].fill = PatternFill("solid", fgColor="1F4788")
        ws_leg.merge_cells("A1:C1")
        ws_leg["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_leg.row_dimensions[1].height = 25

        legend_data = [
            ("A3",  "難度指數（P 值）",       None,   True,  "1F4788"),
            ("A4",  "定義：",                  None,   False, None),
            ("B4",  "該題平均得分 ÷ 題目滿分", None,   False, None),
            ("A5",  "範圍：",                  None,   False, None),
            ("B5",  "0（最難）～ 1（最容易）", None,   False, None),
            ("A6",  "理想範圍：",              None,   False, None),
            ("B6",  "0.25 ～ 0.75（適中難度）",None,   False, None),
            ("A8",  "🟢 容易",                 None,   False, "C6EFCE"),
            ("B8",  "P ≥ 0.75",               None,   False, None),
            ("A9",  "🟡 適中",                 None,   False, "FFEB9C"),
            ("B9",  "0.25 ≤ P < 0.75",        None,   False, None),
            ("A10", "🔴 困難",                 None,   False, "FFC7CE"),
            ("B10", "P < 0.25",               None,   False, None),
            ("A13", "鑑別度（D 值）",          None,   True,  "1F4788"),
            ("A14", "定義：",                  None,   False, None),
            ("B14", "高分組得分率 - 低分組得分率（各取全班 27%）", None, False, None),
            ("A15", "範圍：",                  None,   False, None),
            ("B15", "-1 ～ +1",               None,   False, None),
            ("A17", "⭐ 優良",                 None,   False, "C6EFCE"),
            ("B17", "D ≥ 0.40",               None,   False, None),
            ("A18", "✅ 良好",                 None,   False, "DDEBF7"),
            ("B18", "0.30 ≤ D < 0.40",        None,   False, None),
            ("A19", "⚠️ 尚可",                None,   False, "FFEB9C"),
            ("B19", "0.20 ≤ D < 0.30",        None,   False, None),
            ("A20", "❌ 不佳",                 None,   False, "FFC7CE"),
            ("B20", "D < 0.20（建議檢視此題）",None,   False, None),
            ("A23", "試題品質綜合評價",        None,   True,  "1F4788"),
            ("A24", "最優"),  ("B24", "P 在 0.25～0.75 且 D ≥ 0.40"),
            ("A25", "良好"),  ("B25", "P 在 0.25～0.75 且 D ≥ 0.20"),
            ("A26", "需檢視"), ("B26", "P < 0.25 且 D < 0.20（太難又無鑑別度）"),
            ("A27", "需檢視"), ("B27", "P > 0.75 且 D < 0.20（太易又無鑑別度）"),
        ]
        for entry in legend_data:
            cell_ref = entry[0]
            val      = entry[1]
            is_bold  = entry[3] if len(entry) > 3 else False
            bg_color = entry[4] if len(entry) > 4 else None
            ws_leg[cell_ref] = val
            ws_leg[cell_ref].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws_leg[cell_ref].border = thin_border
            if is_bold:
                ws_leg[cell_ref].font  = Font(bold=True, size=12, color="FFFFFF")
                ws_leg[cell_ref].fill  = PatternFill("solid", fgColor="1F4788")
            if bg_color and not is_bold:
                ws_leg[cell_ref].fill = PatternFill("solid", fgColor=bg_color)
        ws_leg.column_dimensions["A"].width = 22
        ws_leg.column_dimensions["B"].width = 52

        # 格式化其他工作表
        for ws in wb.worksheets:
            if "說明" in ws.title:
                continue
            for cell in ws[1]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border    = thin_border
            ws.row_dimensions[1].height = 25
            ws.freeze_panes = "A2"
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border    = thin_border
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

        # 試題分析顏色
        ws_item = wb["1_試題分析"]
        hdrs = [c.value for c in ws_item[1]]
        dcol = hdrs.index("難度評級") + 1 if "難度評級" in hdrs else None
        kcol = hdrs.index("鑑別評級") + 1 if "鑑別評級" in hdrs else None
        for row in ws_item.iter_rows(min_row=2):
            for cell in row:
                if cell.column in (dcol, kcol):
                    for key, fill in color_fills.items():
                        if key in str(cell.value or ""):
                            cell.fill = fill; break

        # 大題分析顏色
        ws_grp = wb["2_大題分析"]
        hdrs2  = [c.value for c in ws_grp[1]]
        if "難度評級" in hdrs2:
            gc = hdrs2.index("難度評級") + 1
            for row in ws_grp.iter_rows(min_row=2):
                for cell in row:
                    if cell.column == gc:
                        for key, fill in color_fills.items():
                            if key in str(cell.value or ""):
                                cell.fill = fill; break

    print(f"   ✅ Excel 報告已儲存：{output_path}")


# ============================================================
# 主程式
# ============================================================
if __name__ == "__main__":

    exam_info = get_exam_info()

    INPUT_FILE  = "scores.xlsx"
    OUTPUT_FILE = f"{exam_info['file_prefix']}_analysis.xlsx"
    CHART_DIR   = f"{exam_info['file_prefix']}_charts"
    EXAM_TITLE  = exam_info["exam_title"]
    os.makedirs(CHART_DIR, exist_ok=True)

    print(f"\n🔍 載入資料：{INPUT_FILE}")
    df, max_scores, absent_set, paper_map = load_data(INPUT_FILE)
    print(f"   → {len(df)} 名學生，{len(df.columns)} 道題目")

    # ── 試卷加權設定 ──
    # 優先以 Excel 試卷行的分組為準，再套用使用者輸入的比例
    paper_weights = exam_info.get("paper_weights", {"P1": 1.0})
    num_papers    = exam_info.get("num_papers", 1)

    # 若 Excel 試卷行偵測到的試卷數 > 1，但 exam_info 仍是單試卷，自動補齊等權
    papers_in_excel = sorted(set(paper_map.values()))
    if len(papers_in_excel) > 1 and num_papers == 1:
        print(f"   ⚠️  試卷行偵測到 {papers_in_excel}，但設定為單試卷")
        print(f"   ⚠️  已自動改為平均分配各卷比例，建議重新執行並設定正確比例")
        equal_w = round(1.0 / len(papers_in_excel), 4)
        paper_weights = {p: equal_w for p in papers_in_excel}
        num_papers    = len(papers_in_excel)

    # 計算加權總分與各卷得分率
    weighted_scores, paper_pct, weighted_max = calc_weighted_scores(
        df, max_scores, paper_weights, paper_map)

    if num_papers > 1:
        groups = get_paper_groups(df.columns.tolist(), paper_weights, paper_map)
        print(f"   試卷分組：")
        for p, cols in groups.items():
            print(f"   · {p}（{int(paper_weights[p]*100)}%）：{len(cols)} 題 → {', '.join(str(c) for c in cols)}")

    print("\n📐 計算試題分析指標...")
    item_df = item_analysis(df.copy(), max_scores, absent_set)

    print("👨\u200d🎓 計算學生成績摘要...")
    student_df, stats_df = student_summary(
        df.copy(), max_scores, exam_info.get("pass_rate", 0.4), absent_set,
        paper_weights=paper_weights,
        paper_pct=paper_pct,
        weighted_scores=weighted_scores,
        num_papers=num_papers
    )

    print("📋 計算大題分析...")
    group_df = question_group_analysis(df.copy(), max_scores, item_df)

    print("📊 生成視覺化圖表...")
    create_charts(df.copy(), max_scores, item_df.copy(), student_df.copy(), EXAM_TITLE, CHART_DIR, absent_set)

    print("💾 匯出 Excel 報告...")
    export_excel(item_df, group_df, student_df, stats_df, OUTPUT_FILE, EXAM_TITLE)

    # 詢問是否生成個人報告
    gen = input("\n🎓 是否生成每位學生的個人報告？(Y/N)：").strip().upper()
    if gen == "Y":
        try:
            from individual_report import generate_all_reports, generate_combined_class_report

            class_info_raw = pd.read_excel(INPUT_FILE, header=None)
            ci = pd.DataFrame()
            ci["班別"]    = class_info_raw.iloc[4:, 0].values
            ci["班號"]    = class_info_raw.iloc[4:, 1].values
            ci["中文姓名"] = class_info_raw.iloc[4:, 3].values
            ci = ci[ci["中文姓名"].notna() &
                    ~ci["中文姓名"].astype(str).str.contains("說明|輸入", na=False)]

            report_data, rdir, merged_pdf_path = generate_all_reports(
                df, max_scores, item_df, exam_info, ci,
                f"{exam_info['file_prefix']}_個人報告",
                exam_info.get("pass_rate", 0.4),
                absent_set=absent_set
            )

            # 生成全班整合報告
            print(f"\n📄 生成全班整合報告...")
            combined_report_path = generate_combined_class_report(
                df, max_scores, item_df, exam_info, ci,
                f"{exam_info['file_prefix']}_個人報告",
                exam_info.get("pass_rate", 0.4),
                absent_set=absent_set
            )


            if merged_pdf_path:
                print(f"\n✅ PDF 合併完成：{os.path.basename(merged_pdf_path)}")

            print(f"""
╔════════════════════════════════════════════════════╗
║        🎉 個人報告生成完成！                       ║
╠════════════════════════════════════════════════════╣
║  ✅ 共生成 {len(report_data)} 份個人報告              ║
║  📄 個別 Word 檔：(姓名)_個人報告.docx             ║
║  📄 個別 PDF 檔：(姓名)_個人報告.pdf              ║
║  📋 統整 Word：全體學生_個人報告統整.docx        ║
║  📋 統整 PDF：全體學生_個人報告統整.pdf           ║
║  📁 儲存位置：{rdir}/           ║
╚════════════════════════════════════════════════════╝
            """)

        except ImportError as e:
            print(f"\n⚠️  缺少依賴模組")
            print(f"   請執行：pip install python-docx")
            print(f"   詳細：{e}")
        except Exception as e:
            import traceback
            print(f"\n❌ 生成報告時出錯：{e}")
            traceback.print_exc()
    else:
        print("\n✅ 已跳過個人報告生成。")